# import random
# from openpyxl import Workbook

# # -------- Display Function --------
# def display_invoice(products, grand_total, name, num, Mobilenumber):
#     print("\n----------------Invoice Items----------------")
#     print("Customer Name:", name)
#     print("Invoice Number:", num)
#     print("Phone Number:", Mobilenumber)
#     print("\nProduct Name\tProduct Price\tProduct Quantity\tItem Total")
    
#     for item in products:
#         print(f"{item['Product Name']}\t\t{item['Product Price']}\t\t{item['Product Quantity']}\t\t\t{item['Item Total']}")
    
#     print("\n--------------------------------------")
#     print("Grand Total:", grand_total)
#     print("-----------------------------------------")


# # -------- Excel Save Function --------
# def save_to_excel(products, grand_total, name, num, Mobilenumber):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Invoice"

#     # Customer Details
#     ws.append(["Customer Name", name])
#     ws.append(["Invoice Number", num])
#     ws.append(["Phone Number", Mobilenumber])
#     ws.append([])

#     # Table heading
#     ws.append(["Product Name", "Product Price", "Product Quantity", "Item Total"])

#     # Product rows
#     for item in products:
#         ws.append([
#             item["Product Name"],
#             item["Product Price"],
#             item["Product Quantity"],
#             item["Item Total"]
#         ])

#     # Grand total
#     ws.append([])
#     ws.append(["Grand Total", "", "", grand_total])

#     filename = f"Invoice_{num}.xlsx"
#     wb.save(filename)
#     print("\nExcel file saved as:", filename)


# # -------- Main Program --------
# num = random.randint(1000, 9999)
# name = input("Enter Your Name: ")

# Mobilenumber = input("Enter Number: ")
# while not Mobilenumber.isdigit() or len(Mobilenumber) != 10 or Mobilenumber[0] not in ['7','8','9']:
#     print("Enter a valid 10-digit mobile number")
#     Mobilenumber = input("Enter Number: ")

# products = []
# grand_total = 0

# while True:
#     productname = input("Enter Product Name: ")

#     productprice = input("Enter Product Price: ")
#     while not productprice.isdigit():
#         print("Enter a valid number")
#         productprice = input("Enter Product Price: ")

#     productquantiti = input("Enter Product Quantity: ")
#     while not productquantiti.isdigit():
#         print("Enter a valid number")
#         productquantiti = input("Enter Product Quantity: ")

#     item_total = int(productprice) * int(productquantiti)
#     grand_total += item_total

#     products.append({
#         "Product Name": productname,
#         "Product Price": int(productprice),
#         "Product Quantity": int(productquantiti),
#         "Item Total": item_total
#     })

#     while True:
#         print("\n1. Add another item")
#         print("2. Remove an item")
#         print("3. Update item")
#         print("4. Display Invoice")
#         print("5. Exit")
#         choice = input("Enter Your choice: ")

#         if choice == '1':
#             break  

#         elif choice == '2':
#             remove_item = input("Enter the name of the item to remove: ")
#             for item in products:
#                 if item["Product Name"].lower() == remove_item.lower():
#                     grand_total -= item["Item Total"]
#                     products.remove(item)
#                     print(f"{remove_item} has been removed.")
#                     break
#             else:
#                 print(f"{remove_item} not found in the invoice.")

#         elif choice == '3':
#             update_item = input("Enter the name of the item to update: ")
#             for item in products:
#                 if item["Product Name"].lower() == update_item.lower():
#                     new_price = input("Enter new price: ")
#                     while not new_price.isdigit():
#                         print("Enter a valid number")
#                         new_price = input("Enter new price: ")

#                     new_quantity = input("Enter new quantity: ")
#                     while not new_quantity.isdigit():
#                         print("Enter a valid number")
#                         new_quantity = input("Enter new quantity: ")

#                     old_total = item["Item Total"]

#                     item["Product Price"] = int(new_price)
#                     item["Product Quantity"] = int(new_quantity)
#                     item["Item Total"] = int(new_price) * int(new_quantity)

#                     grand_total += (item["Item Total"] - old_total)

#                     print(f"{update_item} has been updated.")
#                     break
#             else:
#                 print(f"{update_item} not found in the invoice.")

#         elif choice == '4':
#             display_invoice(products, grand_total, name, num, Mobilenumber)

#         elif choice == '5':
#             save_to_excel(products, grand_total, name, num, Mobilenumber)
#             break

#         else:
#             print("Invalid choice. Please enter 1,2,3,4 or 5")

#     if choice == '5':
#         break  

# # Final display
# display_invoice(products, grand_total, name, num, Mobilenumber)


import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# -------- Display Function --------
def display_invoice(products, grand_total, name, num, Mobilenumber):
    print("\n----------------Invoice Items----------------")
    print("Customer Name:", name)
    print("Invoice Number:", num)
    print("Phone Number:", Mobilenumber)
    print("\nProduct Name\tProduct Price\tProduct Quantity\tItem Total")
    
    for item in products:
        print(f"{item['Product Name']}\t\t{item['Product Price']}\t\t{item['Product Quantity']}\t\t\t{item['Item Total']}")
    
    print("\n--------------------------------------")
    print("Grand Total:", grand_total)
    print("-----------------------------------------")


# -------- Excel Save Function with Formatting --------
def save_to_excel(products, grand_total, name, num, Mobilenumber):
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"), 
        right=Side(style="thin"), 
        top=Side(style="thin"), 
        bottom=Side(style="thin")
    )

    # Customer Details
    ws.append(["Customer Name", name])
    ws.append(["Invoice Number", num])
    ws.append(["Phone Number", Mobilenumber])
    ws.append([])

    # Table heading
    headers = ["Product Name", "Product Price", "Product Quantity", "Item Total"]
    ws.append(headers)

    for col in range(1, 5):
        ws.cell(row=5, column=col).font = bold_font
        ws.cell(row=5, column=col).alignment = center_align
        ws.cell(row=5, column=col).border = thin_border

    # Product rows
    for i, item in enumerate(products, start=6):
        ws.append([
            item["Product Name"],
            item["Product Price"],
            item["Product Quantity"],
            item["Item Total"]
        ])
        for col in range(1, 5):
            ws.cell(row=i, column=col).border = thin_border
            ws.cell(row=i, column=col).alignment = center_align

    # Grand total
    ws.append([])
    ws.append(["Grand Total", "", "", grand_total])
    ws.cell(row=ws.max_row, column=1).font = bold_font
    ws.cell(row=ws.max_row, column=4).font = bold_font
    ws.cell(row=ws.max_row, column=1).alignment = center_align
    ws.cell(row=ws.max_row, column=4).alignment = center_align

    # Adjust column widths
    column_widths = [20, 15, 18, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    filename = f"Invoice_{num}.xlsx"
    wb.save(filename)
    print("\nExcel file saved as:", filename)


# -------- Main Program --------
num = random.randint(1000, 9999)
name = input("Enter Your Name: ")

Mobilenumber = input("Enter Number: ")
while not Mobilenumber.isdigit() or len(Mobilenumber) != 10 or Mobilenumber[0] not in ['7','8','9']:
    print("Enter a valid 10-digit mobile number")
    Mobilenumber = input("Enter Number: ")

products = []
grand_total = 0

while True:
    productname = input("Enter Product Name: ")

    productprice = input("Enter Product Price: ")
    while not productprice.isdigit():
        print("Enter a valid number")
        productprice = input("Enter Product Price: ")

    productquantiti = input("Enter Product Quantity: ")
    while not productquantiti.isdigit():
        print("Enter a valid number")
        productquantiti = input("Enter Product Quantity: ")

    item_total = int(productprice) * int(productquantiti)
    grand_total += item_total

    products.append({
        "Product Name": productname,
        "Product Price": int(productprice),
        "Product Quantity": int(productquantiti),
        "Item Total": item_total
    })

    while True:
        print("\n1. Add another item")
        print("2. Remove an item")
        print("3. Update item")
        print("4. Display Invoice")
        print("5. Save & Exit")
        choice = input("Enter Your choice: ")

        if choice == '1':
            break  

        elif choice == '2':
            remove_item = input("Enter the name of the item to remove: ")
            for item in products:
                if item["Product Name"].lower() == remove_item.lower():
                    grand_total -= item["Item Total"]
                    products.remove(item)
                    print(f"{remove_item} has been removed.")
                    break
            else:
                print(f"{remove_item} not found in the invoice.")

        elif choice == '3':
            update_item = input("Enter the name of the item to update: ")
            for item in products:
                if item["Product Name"].lower() == update_item.lower():
                    new_price = input("Enter new price: ")
                    while not new_price.isdigit():
                        print("Enter a valid number")
                        new_price = input("Enter new price: ")

                    new_quantity = input("Enter new quantity: ")
                    while not new_quantity.isdigit():
                        print("Enter a valid number")
                        new_quantity = input("Enter new quantity: ")

                    old_total = item["Item Total"]

                    item["Product Price"] = int(new_price)
                    item["Product Quantity"] = int(new_quantity)
                    item["Item Total"] = int(new_price) * int(new_quantity)

                    grand_total += (item["Item Total"] - old_total)

                    print(f"{update_item} has been updated.")
                    break
            else:
                print(f"{update_item} not found in the invoice.")

        elif choice == '4':
            display_invoice(products, grand_total, name, num, Mobilenumber)

        elif choice == '5':
            save_to_excel(products, grand_total, name, num, Mobilenumber)
            break

        else:
            print("Invalid choice. Please enter 1,2,3,4 or 5")

    if choice == '5':
        break  

# Final display
display_invoice(products, grand_total, name, num, Mobilenumber)

